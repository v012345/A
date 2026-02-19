import json
import sys
import os
import subprocess

try:
    import pandas as pd
except:
    subprocess.Popen([sys.executable, "-m", "pip", "install", "pandas"],
                     creationflags=subprocess.CREATE_NEW_CONSOLE).wait()
    import pandas as pd

try:
    import openpyxl
except:
    subprocess.Popen([sys.executable, "-m", "pip", "install", "openpyxl"],
                     creationflags=subprocess.CREATE_NEW_CONSOLE).wait()
    import openpyxl


from optparse import OptionParser
parser = OptionParser()
parser.add_option("--xlsx", action="store",
                  dest="xlsx", type="string", help="设计书表格路径")
(opts, args) = parser.parse_args()

# ===== 状态机读入数据, 使用关键字完全匹配改变状态 =====

enter_parse_table_state_keywords = [
    "取得テーブル"
]

def parse_table_to_json(xtable):
    result = {
        "table": None,
        "joins": [],
        "where": {}
    }
    result["table"] = xtable.iloc[xtable.row+1, xtable.col]  # 主表

    table_name_col = xtable.col
    
    join_type_col = xtable.find_str_in_row(xtable.row, "結合条件")
    where_row = xtable.find_next_str_in_col(xtable.row, xtable.col, "検索条件")
    
    # 目前这个例子太小了
    # 如果没有 "検索条件", 那么会不知到怎么退出函数
    # 针对这个例子来说, 现在的实现没有问题
    
    # 查看是否有 "結合条件"
    if join_type_col != -1:
        for row in range(xtable.row + 2, where_row): # 这里的 where_row 要优化, 目前只能处理 "検索条件" 在 "結合条件" 下方的情况
            join = {
                "type": None,
                "table": None,
                "on": None
            }
            join["type"] = xtable.iloc[row, join_type_col]
            join["table"] = xtable.iloc[row, table_name_col]
            join["on"] = xtable.iloc[row, join_type_col + 1]  # 假设 "結合条件" 的下一列是 ON 条件
            result["joins"].append(join)
  
    if where_row != -1:
        end_row = xtable.find_empty_in_col(where_row + 1, table_name_col + 1)  # 寻找 "検索条件" 列的第一个空行, 以确定 where 条件的范围
        if end_row == -1:
            end_row = xtable.max_row
        # 例子里只有 AND 条件, 没有 OR 条件和括号, 所以暂时不考虑复杂的 where 条件
        where = {
            "logic": "AND",  # 目前默认都是 AND 条件
            "conditions": []
        }
        for row in range(where_row + 1, end_row):
            # 可以改成配置文件, 目前先写死, 主要就是偏移量的问题
           condition = {
                "field": xtable.iloc[row, table_name_col + 1],  # 假设 "検索条件" 的下一列是字段名
                "operator": xtable.iloc[row, table_name_col + 4],  # 假设 "検索条件" 的下一列的下一列是操作符
                "value": xtable.iloc[row, table_name_col + 5]  # 假设 "検索条件" 的下一列的下一列的下一列是值
           }
           where["conditions"].append(condition)
        result["where"] = where

    return result


class XlsxTable:
    def __init__(self, df):
        rows, cols = df.shape
        self.row = 0
        self.col = 0
        self.max_row = rows
        self.max_col = cols
        self.iloc = df.iloc
    def next(self):
        self.col += 1
        if self.col >= self.max_col:
            self.col = 0
            self.row += 1
        if self.row >= self.max_row:
            raise StopIteration
    def __str__(self):
        return f"({self.row}, {self.col})"
    
    def find_str_in_row(self, row, s):
        for c in range(self.max_col):
            if str(self.iloc[row, c]) == s:
                return c
        return -1
    
    def find_next_str_in_col(self,row, col, s):
        for r in range(row, self.max_row):
            if str(self.iloc[r, col]) == s:
                return r
        return -1
    
    def find_empty_in_col(self, row, col):
        for r in range(row, self.max_row):
            value = self.iloc[r, col]
            if value is None or pd.isna(value):
                return r
        return -1

if __name__ == '__main__':
    print(f"处理表 : {opts.xlsx}")
    xlsx_file = opts.xlsx
    if not os.path.exists(xlsx_file):
        xlsx_file = os.path.join(os.getcwd(), xlsx_file)
        if not os.path.exists(xlsx_file):
            print("文件不存在")
            sys.exit(1)
    df = pd.read_excel(xlsx_file, engine="openpyxl", header=None)
    result = []
    xlsx = XlsxTable(df)
    try:
        while True:
            value = xlsx.iloc[xlsx.row, xlsx.col]
            if str(value) in enter_parse_table_state_keywords:
                temp_row, temp_col = xlsx.row, xlsx.col
                result.append(parse_table_to_json(xlsx))
                xlsx.row, xlsx.col = temp_row, temp_col  # 防止随便粘贴表格
                xlsx.next() 
            else:
                xlsx.next()
    except StopIteration:
        print("遍历完成")
    forAI = {
        "cmd": "Convert the SQL into ABAP code.",
        "sql": result
    }
    with open("forAI.json", "w", encoding="utf-8") as f:
        json.dump(forAI, f, ensure_ascii=False, indent=4)
    print("生成 forAI.json 完成")